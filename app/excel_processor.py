"""Excel/CSV analytics engine.

Pipeline:
1. Parse uploaded file → extract schema, dtypes, sample rows, summary stats
2. Convert to Parquet for optimized querying
3. LLM receives ONLY schema + metadata (never full data)
4. LLM generates structured analytics JSON
5. DuckDB executes the analytics safely
6. Results returned with chart config for frontend rendering
"""

import hashlib
import json
import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    import polars as pl
except ImportError:
    pl = None  # type: ignore

try:
    import duckdb
except ImportError:
    duckdb = None  # type: ignore

# SQL keywords that are NEVER allowed
BLOCKED_SQL = re.compile(
    r"\b(DROP|DELETE|UPDATE|INSERT|ALTER|CREATE|TRUNCATE|EXEC|EXECUTE|GRANT|REVOKE|MERGE)\b",
    re.IGNORECASE,
)


def _file_id(filepath: Path) -> str:
    raw = f"{filepath.name}:{filepath.stat().st_size}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


def parse_spreadsheet(filepath: Path) -> dict[str, Any]:
    """Parse an Excel/CSV file and extract schema + sample data.

    Returns metadata dict (never full data).
    """
    if pl is None:
        raise RuntimeError("polars is required. Run: pip install polars")

    suffix = filepath.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel files. Run: pip install openpyxl")
        df = pl.read_excel(filepath)
    elif suffix == ".csv":
        df = pl.read_csv(filepath, infer_schema_length=1000)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    file_id = _file_id(filepath)

    # Extract schema
    schema = []
    for col in df.columns:
        dtype = str(df[col].dtype)
        non_null = df[col].drop_nulls().len()
        schema.append({
            "column": col,
            "dtype": dtype,
            "non_null_count": non_null,
            "null_count": len(df) - non_null,
        })

    # Summary statistics for numeric columns
    stats = {}
    numeric_cols = [col for col in df.columns if df[col].dtype in (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8, pl.UInt64, pl.UInt32)]
    for col in numeric_cols:
        series = df[col].drop_nulls()
        if len(series) == 0:
            continue
        stats[col] = {
            "min": float(series.min()),  # type: ignore
            "max": float(series.max()),  # type: ignore
            "mean": float(series.mean()),  # type: ignore
            "median": float(series.median()),  # type: ignore
            "std": float(series.std()) if len(series) > 1 else 0.0,  # type: ignore
        }

    # Sample rows (max 5)
    sample_rows = df.head(5).to_dicts()

    return {
        "file_id": file_id,
        "filename": filepath.name,
        "row_count": len(df),
        "column_count": len(df.columns),
        "columns": df.columns,
        "schema": schema,
        "statistics": stats,
        "sample_rows": sample_rows,
    }


def convert_to_parquet(filepath: Path, output_dir: Path) -> Path:
    """Convert Excel/CSV to Parquet for optimized querying."""
    if pl is None:
        raise RuntimeError("polars is required")

    suffix = filepath.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        df = pl.read_excel(filepath)
    elif suffix == ".csv":
        df = pl.read_csv(filepath, infer_schema_length=1000)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    output_dir.mkdir(parents=True, exist_ok=True)
    parquet_path = output_dir / f"{filepath.stem}.parquet"
    df.write_parquet(parquet_path)
    logger.info("Converted %s to parquet: %s (%d rows)", filepath.name, parquet_path, len(df))
    return parquet_path


def validate_sql(sql: str) -> tuple[bool, str]:
    """Validate that generated SQL is safe to execute."""
    if BLOCKED_SQL.search(sql):
        return False, f"Blocked SQL keyword found in: {sql[:100]}"
    if ";" in sql and sql.strip().count(";") > 1:
        return False, "Multiple statements not allowed"
    return True, "ok"


def execute_analytics(parquet_path: Path, sql: str) -> dict[str, Any]:
    """Execute a SQL query against a parquet file using DuckDB.

    Args:
        parquet_path: Path to the parquet file.
        sql: A SELECT query. The table name in SQL should be 'data'.

    Returns:
        {columns: [...], rows: [...], row_count: int}
    """
    if duckdb is None:
        raise RuntimeError("duckdb is required. Run: pip install duckdb")

    is_safe, reason = validate_sql(sql)
    if not is_safe:
        raise ValueError(f"Unsafe SQL blocked: {reason}")

    conn = duckdb.connect(":memory:")
    try:
        conn.execute(f"CREATE VIEW data AS SELECT * FROM read_parquet('{parquet_path}')")
        result = conn.execute(sql)
        columns = [desc[0] for desc in result.description]
        rows = result.fetchall()

        # Convert to list of dicts
        row_dicts = [dict(zip(columns, row)) for row in rows]

        return {
            "columns": columns,
            "rows": row_dicts,
            "row_count": len(row_dicts),
        }
    finally:
        conn.close()


def execute_to_dataframe(parquet_path: Path, sql: str):
    """Run a validated SELECT and return the result as a Polars DataFrame.

    Used by the export pipeline so we can write the result back to .xlsx
    without round-tripping through Python dicts.
    """
    if duckdb is None:
        raise RuntimeError("duckdb is required. Run: pip install duckdb")
    if pl is None:
        raise RuntimeError("polars is required. Run: pip install polars")

    is_safe, reason = validate_sql(sql)
    if not is_safe:
        raise ValueError(f"Unsafe SQL blocked: {reason}")

    conn = duckdb.connect(":memory:")
    try:
        conn.execute(f"CREATE VIEW data AS SELECT * FROM read_parquet('{parquet_path}')")
        # DuckDB → Arrow → Polars is zero-copy and preserves dtypes.
        arrow_tbl = conn.execute(sql).fetch_arrow_table()
        return pl.from_arrow(arrow_tbl)
    finally:
        conn.close()


def export_query_to_excel(
    parquet_path: Path,
    sql: str,
    output_path: Path,
    source_xlsx: Path | None = None,
) -> dict[str, Any]:
    """Run a SELECT against the parquet and persist a **styled** .xlsx.

    Styling applied:
        - Bold white header row on a dark slate fill, frozen.
        - Subtle row borders so the table reads as a table.
        - Auto-fit column widths (clamped between 10 and 50 chars).
        - Smart number formats: ints get thousand-separators, floats
          get 2-decimal thousand-separators, datetimes/dates get an
          ISO format.
        - If ``source_xlsx`` is an existing Excel file, each column's
          original ``number_format`` (currency, percentage, custom date
          masks, etc.) is reused for matching column names so the export
          looks like a filtered view of the original — not a raw dump.

    Returns ``{row_count, column_count, columns, path}``.
    """
    if pl is None:
        raise RuntimeError("polars is required")

    df = execute_to_dataframe(parquet_path, sql)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    original_formats: dict[str, str] = {}
    if source_xlsx and source_xlsx.exists() and source_xlsx.suffix.lower() in (".xlsx", ".xls"):
        original_formats = _read_original_column_formats(source_xlsx)

    _write_styled_workbook(df, output_path, original_formats)

    return {
        "row_count": len(df),
        "column_count": len(df.columns),
        "columns": list(df.columns),
        "path": str(output_path),
    }


def _read_original_column_formats(xlsx_path: Path) -> dict[str, str]:
    """Read each column's ``number_format`` from the first data row.

    Returns a mapping of column header → openpyxl number_format code
    (e.g. ``"$#,##0.00"``, ``"yyyy-mm-dd"``). Columns with the default
    ``"General"`` format are skipped so we don't override smarter defaults.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=False)
        try:
            ws = wb.active
            if ws is None:
                return {}
            rows_iter = ws.iter_rows(min_row=1, max_row=2)
            header_row = next(rows_iter, None)
            sample_row = next(rows_iter, None)
            if not header_row or not sample_row:
                return {}
            formats: dict[str, str] = {}
            for header_cell, sample_cell in zip(header_row, sample_row):
                name = header_cell.value
                fmt = sample_cell.number_format
                if name is None or not fmt or fmt == "General":
                    continue
                formats[str(name)] = fmt
            return formats
        finally:
            wb.close()
    except Exception as e:
        logger.warning("Could not read original column formats from %s: %s", xlsx_path, e)
        return {}


def _write_styled_workbook(df, output_path: Path, original_formats: dict[str, str]) -> None:
    """Write a Polars DataFrame to ``output_path`` with professional styling."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for styled export. Run: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "result"

    columns = list(df.columns)

    # ── styles ────────────────────────────────────────────────────
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", indent=1)
    thin = Side(border_style="thin", color="E5E7EB")
    row_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── header row ────────────────────────────────────────────────
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = row_border
    ws.row_dimensions[1].height = 24

    # ── data rows ─────────────────────────────────────────────────
    # Cache per-column formatting decisions so we don't recompute per cell.
    col_formats: dict[str, str | None] = {}
    for col in columns:
        if col in original_formats:
            col_formats[col] = original_formats[col]
            continue
        dtype = df.schema.get(col)
        col_formats[col] = _default_number_format(dtype)

    for row_idx, row in enumerate(df.iter_rows(named=True), start=2):
        for col_idx, col in enumerate(columns, start=1):
            v = row[col]
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            fmt = col_formats.get(col)
            if fmt:
                cell.number_format = fmt
            cell.border = row_border

    # ── auto-fit column widths ────────────────────────────────────
    # Sample the first 200 rows so we don't iterate the whole DataFrame
    # for a width heuristic on huge exports.
    sample_n = min(len(df), 200)
    head = df.head(sample_n)
    for col_idx, col in enumerate(columns, start=1):
        max_len = len(str(col))
        if sample_n > 0:
            for v in head[col].to_list():
                if v is None:
                    continue
                w = len(_render_for_width(v))
                if w > max_len:
                    max_len = w
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 3, 50), 10)

    # ── freeze header + filter ────────────────────────────────────
    ws.freeze_panes = "A2"
    if len(df) > 0:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    wb.save(output_path)


def _default_number_format(dtype) -> str | None:
    """Pick a sensible openpyxl number_format for a Polars dtype."""
    if pl is None or dtype is None:
        return None
    # Integer-like
    if dtype in (pl.Int64, pl.Int32, pl.Int16, pl.Int8, pl.UInt64, pl.UInt32, pl.UInt16, pl.UInt8):
        return "#,##0"
    # Float-like
    if dtype in (pl.Float64, pl.Float32):
        return "#,##0.00"
    # Date / datetime
    if dtype == pl.Date:
        return "yyyy-mm-dd"
    if dtype == pl.Datetime:
        return "yyyy-mm-dd hh:mm:ss"
    return None


def _render_for_width(v: Any) -> str:
    """Approximate how wide a value will appear in Excel for width sizing."""
    if isinstance(v, float):
        # 2-decimal thousand-grouped renders longer than repr()
        return f"{v:,.2f}"
    if isinstance(v, int):
        return f"{v:,}"
    return str(v)


def select_chart_type(columns: list[str], rows: list[dict], query: str = "") -> dict[str, Any]:
    """Automatically select the best chart type based on data shape and query intent.

    Returns a chart configuration dict for the frontend.
    Multi-series charts are returned when multiple numeric columns exist.
    """
    if not rows or not columns:
        return {"type": "none", "reason": "No data to visualize"}

    num_cols = []
    cat_cols = []
    for col in columns:
        sample = rows[0].get(col)
        if isinstance(sample, (int, float)):
            num_cols.append(col)
        else:
            cat_cols.append(col)

    query_lower = query.lower()
    x_col = cat_cols[0] if cat_cols else columns[0]

    # Multi-series detection: if we have >1 numeric col + at least 1 categorical, group them
    has_multi_series = len(num_cols) >= 2 and len(cat_cols) >= 1
    series = [{"key": c, "name": c} for c in num_cols] if has_multi_series else []

    # Area chart
    if any(kw in query_lower for kw in ["area", "filled", "cumulative"]):
        return {
            "type": "area",
            "x": x_col,
            "series": series if series else [{"key": num_cols[0], "name": num_cols[0]}],
            "title": "Area Analysis",
        }

    # Time-series / trend detection
    time_keywords = ["trend", "over time", "monthly", "yearly", "daily", "time series", "growth"]
    if any(kw in query_lower for kw in time_keywords):
        return {
            "type": "line",
            "x": x_col,
            "series": series if series else [{"key": num_cols[0], "name": num_cols[0]}],
            "title": "Trend Analysis",
        }

    # Proportion detection
    proportion_keywords = ["proportion", "percentage", "share", "distribution", "breakdown", "pie"]
    if any(kw in query_lower for kw in proportion_keywords) and len(rows) <= 10:
        return {
            "type": "pie",
            "labels": x_col,
            "values": num_cols[0] if num_cols else columns[-1],
            "title": "Distribution",
        }

    # Correlation / scatter
    if len(num_cols) >= 2 and any(kw in query_lower for kw in ["correlation", "scatter", "relationship"]):
        return {
            "type": "scatter",
            "x": num_cols[0],
            "y": num_cols[1],
            "title": "Correlation",
        }

    # Radar / spider chart
    if any(kw in query_lower for kw in ["radar", "spider", "profile", "compare across"]):
        return {
            "type": "radar",
            "x": x_col,
            "series": series if series else [{"key": num_cols[0], "name": num_cols[0]}],
            "title": "Radar Comparison",
        }

    # Stacked bar
    if any(kw in query_lower for kw in ["stacked", "stack", "total breakdown"]):
        return {
            "type": "stacked_bar",
            "x": x_col,
            "series": series if series else [{"key": num_cols[0], "name": num_cols[0]}],
            "title": "Stacked Comparison",
        }

    # Histogram
    if len(num_cols) == 1 and len(cat_cols) == 0:
        return {
            "type": "histogram",
            "x": num_cols[0],
            "title": "Distribution",
        }

    # Default: grouped bar (multi-series when possible)
    return {
        "type": "bar",
        "x": x_col,
        "series": series if series else [{"key": num_cols[0] if num_cols else columns[-1], "name": num_cols[0] if num_cols else "Value"}],
        "title": "Comparison",
    }
